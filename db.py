import MySQLdb
import time
import openpyxl

from contextlib import contextmanager

@contextmanager
def create_connection():
    # Connect to the database
    conn = MySQLdb.connect(
        host="eyobworku.mysql.pythonanywhere-services.com",
        user="eyobworku",
        password="12345ewp",
        db="eyobworku$wkuList"
    )
    try:
        yield conn
    finally:
        conn.close()

def insert_users(chat_id, roll,department,class_id=None,name=None):
    with create_connection() as conn:
        cursor = conn.cursor()
        query = "INSERT INTO users (chat_id, class_id,name, roll,department) VALUES (%s, %s,%s,%s,%s)"
        values = (chat_id,class_id,name, roll,department)
        cursor.execute(query, values)
        print(values)
    # Commit the changes
        conn.commit()
        if cursor.rowcount > 0:
            return True
        else:
            return False

def select_user(chat_id):
    with create_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM users where chat_id={chat_id}")
        result_set = cursor.fetchall()
        column_names = [desc[0] for desc in cursor.description]

        row_dict = {}
        for row in result_set:
            for i, column_value in enumerate(row):
                column_name = column_names[i]
                row_dict[column_name] = column_value
    return row_dict

def upate_user(chat_id, roll,department,class_id=None,name=None):
    with create_connection() as conn:
        cursor = conn.cursor() #UPDATE mytable SET department=%s, class_id=%s WHERE roll=%s
        query = "UPDATE users set class_id=%s ,name=%s, roll=%s, department=%s WHERE chat_id=%s"
        values = (class_id,name, roll,department,chat_id)
        cursor.execute(query, values)
    # Commit the changes
        conn.commit()
        if cursor.rowcount == 1:
            return True
        else:
            return False

def delete_user(chat_id):
    with create_connection() as conn:
        cursor = conn.cursor()
        query = "DELETE FROM users WHERE chat_id=%s"
        value = (chat_id,)
        cursor.execute(query,value)
        conn.commit()
        if cursor.rowcount > 0:
            return True
        else:
            return False
def readJsonXl(file):
    timestamp = int(time.time())
    timestamp_str = "{:08d}".format(timestamp)
    fileSlah = file.split('/')[-1]
    fileDot = fileSlah.split('.')[0]
    fileName = fileDot+timestamp_str
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    keys = [str(ws.cell(row=1,column=i).value) for i in range(1,ws.max_column+1)]

    values = list()
    for value in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column,values_only=True):
        values.append(value)
    return fileName, keys, values

def create_and_insert_table(file):
    # call the readJsonXl function to get the filename, keys, and values
    fileName, keys, values = readJsonXl(file)

    # use the keys list to create the table
    try:
        with create_connection() as conn:
            cursor = conn.cursor()
            create_table_query = "CREATE TABLE `{}` ({})".format(fileName, ", ".join(["`{}` VARCHAR(255)".format(key) for key in keys]))
            cursor.execute(create_table_query)
            conn.commit()
            # insert the values into the table
            for value in values:
                # insert_query = "INSERT INTO {} ({}) VALUES ({})".format(fileName, ", ".join(keys), ", ".join(["%s" for i in range(len(keys))]))
                insert_query = "INSERT INTO `{}` ({}) VALUES ({})".format(fileName, ", ".join(["`{}`".format(key) for key in keys]), ", ".join(["%s" for i in range(len(keys))]))
                cursor.execute(insert_query, value)
            conn.commit()
            return fileName, keys, True
    except MySQLdb.Error as e:
        with create_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DROP TABLE {}'.format(fileName))
            print("An error occurred:", e)
            return fileName, keys, False

def insert_reults(chat_id,fileName,keys):
    with create_connection() as conn:
        keys_str = '-'.join(keys)
        department = select_user(chat_id)['department']
        cursor = conn.cursor()
        query = "INSERT INTO resultList (chat_id, department,table_name, keys_value) VALUES (%s, %s, %s, %s)"
        values = (chat_id,department,fileName, keys_str)
        cursor.execute(query, values)
    # Commit the changes
        conn.commit()
        if cursor.rowcount > 0:
            return True
        else:
            return False

def select_results(chat_id):
    user = select_user(chat_id)
    if 'chat_id' in user:
        with create_connection() as conn:
            cursor = conn.cursor()
            dep = user['department']
            query = "SELECT `table_name` FROM `resultList` where department=%s"
            values = (dep,)
            cursor.execute(query,values)
            result_set = cursor.fetchall()
            return result_set
    else:
        return []

def select_keys(table_name,chat_id):
    user = select_user(chat_id)
    if 'chat_id' in user:
        with create_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT `keys_value` FROM resultList where table_name=%s"
            values = (table_name,)
            cursor.execute(query,values)
            result_set = cursor.fetchall()
            keys_res = result_set[0][0]
            keys = keys_res.split('-')

            query = "SELECT * FROM {} where id=%s".format(table_name)
            values = (user['class_id'],)
            cursor.execute(query,values)
            result_set = cursor.fetchall()[0]

            row_dict = {}
            for i, column_value in enumerate(result_set):
                column_name = keys[i]
                row_dict[column_name] = column_value
            return row_dict
    else:
        return {}



