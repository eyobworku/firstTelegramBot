import json
import requests
import os
import openpyxl

BOT_TOKEN = '6003301685:AAFSmqvk4IWe7mn9xt2HsLagLMjV1CIzpJA'
# methods = []
# post_fields = [{}]


def echo_input(update,chat_id=949587899):
        # Using the Telegram Bot API method "sendMessage",
        # https://core.telegram.org/bots/api#sendmessage
        # methods.append('sendMessage')

        # # It's going back to the same chat where it came from
        # post_fields[0]['chat_id'] = update['message']['chat']['id']
        update['methods'].append('sendMessage')
        update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
        update['post_fields'][0]['text'] = json.dumps(update,indent=2)

        # The text being returned is just the whole object
        # post_fields[0]['disable_web_page_preview'] = True

def send_response(update):
        # Now go through the array and send each method and post_fields
        for i in range(len(update['methods'])):
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/{update['methods'][i]}"
            headers = {"Content-Type": "application/json"}
            response = requests.post(url, headers=headers, json=update['post_fields'][i])
            return response.json()
            # Process the response if needed

def save_file(update):
    myFile = {}
    myFile['methods'] = []
    myFile['post_fields'] = [{}]
    myFile['methods'].append('getFile')
    myFile['post_fields'][0]['file_id'] = update['message']['document']['file_id']
    myFile = send_response(myFile)

    url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{myFile['result']['file_path']}"
    response = requests.get(url)
    folder_path = 'savedFile'
    os.makedirs(folder_path, exist_ok=True)

    file_name = myFile['result']['file_path'].split('/')
    file_name = file_name[-1]
    my_file = os.path.join(folder_path,file_name)
    if response.status_code == 200:
        with open(my_file, 'wb') as file:
            file.write(response.content)
    myList = readJsonXl(my_file)

    update['methods'].append('sendMessage')
    update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
    update['post_fields'][0]['text'] = json.dumps(myList,indent=2)


def route_requests(update):
    if 'entities' in update['message'] and update['message']['entities'] and update['message']['entities'][0]['type'] == 'bot_command':
        update['parameters'] = []
        update['command'] = [update['message']['text'][0:update['message']['entities'][0]['length']]]
        update['parameters'].append(update['message']['text'][update['message']['entities'][0]['length']+1:])

        if update['command'][0] == '/whoami':
            # whoami(update)
            update['methods'].append('sendMessage')
            update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
            update['post_fields'][0]['text'] ='who am i'
        elif update['command'][0] == '/echo':
            echo_input(update)
        elif update['command'][0] == '/gethelp':
            update['methods'] = ['sendMessage']
            update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
            update['post_fields'][0]['text'] ='get help'
            # perform_help(update)

        else:
            # bad_request(update)
            update['methods'].append('sendMessage')
            update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
            update['post_fields'][0]['text'] ='bad request'
    elif 'document' in update['message'] and update['message']['document']:
        save_file(update)
    else:
        if 'reply_to_message' in update['message']:
            # perform_reply(update)
            update['methods'].append('sendMessage')
            update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
            update['post_fields'][0]['text'] ='perform reply'
        else:
            # # perform_text(update)
            update['methods'].append('sendMessage')
            update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
            update['post_fields'][0]['text'] ='perform text'
            # echo_input(update)

def readJsonXl(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    keys = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]

    values = list()
    for value in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column,values_only=True):
        values.append(value)
    ##convert
    myList = list()
    for item in values:
        myList.append({k: v for k,v in zip(keys,item)})
    return myList



# def whoami(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']

#     firstname = update['message']['from']['first_name']
#     lastname = update['message']['from']['last_name']
#     username = update['message']['from']['username']
#     user_id = update['message']['from']['id']

#     update['post_fields'][0]['text'] = f"Hello, {firstname} {lastname}!\nYour username is {username} and your user ID is {user_id}."


# def perform_help(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
#     update['post_fields'][0]['text'] = 'What kind of help would you like? Type "tech", "billing", or "other".'

#     keyboard = [
#         [{'text': 'tech'}, {'text': 'billing'}],
#         [{'text': 'other'}],
#     ]
#     update['post_fields'][0]['reply_markup'] = json.dumps({
#         'keyboard': keyboard,
#         'one_time_keyboard': True,
#     })


# def bad_request(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
#     update['post_fields'][0]['text'] = f'You wrote "{update["message"]["text"]}". I don\'t understand that.'
#     update['post_fields'][0]['disable_web_page_preview'] = True


# def perform_reply(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
#     reply_text = update['message']['reply_to_message']['text']
#     if reply_text == 'Please describe your problem and I\'ll forward your message.':
#         finish_reply(update)


# def perform_text(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']

#     text = update['message']['text']
#     if text == 'tech':
#         update['post_fields'][0]['text'] = 'Tech info here.'
#     elif text == 'billing':
#         update['post_fields'][0]['text'] = 'Billing info here.'
#     elif text == 'other':
#         forward_other(update)
#     else:
#         update['post_fields'][0]['text'] = "I didn't understand that."


# def finish_reply(update):
#     update['post_fields'][0]['chat_id'] = CUSTOMER_SERVICE_ID
#     update['post_fields'][0]['text'] = (
#         "You received the following question:\n\n" +
#         update['message']['text'] +
#         "\n\nIt was sent by\n\n" +
#         str(update['message']['from'])
#     )

#     update['method'].append('sendMessage')
#     update['post_fields'].append({})
#     update['post_fields'][1]['chat_id'] = update['message']['chat']['id']
#     update['post_fields'][1]['text'] = 'Your help request has been forwarded. You should get a response within 24 hours.'


# def forward_other(update):
#     update['method'] = ['sendMessage']
#     update['post_fields'][0]['chat_id'] = update['message']['chat']['id']
#     update['post_fields'][0]['text'] = "Please describe your problem and I'll forward your message."
#     update['post_fields'][0]['reply_markup'] = json.dumps({
#         'force_reply': True,
#     })